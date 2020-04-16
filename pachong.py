import requests
import os
from tktest import *
from bs4 import BeautifulSoup
from openpyxl import Workbook
import openpyxl


def get_information(url):
    # url = "http://www.ygsoft.com/BigData/index.html?from=baidu"
    page = requests.get(url)
    print(page.status_code)
    print(page.content)

    soup = BeautifulSoup(page.content, 'html.parser')

    print(soup.prettify())

    name = soup.find_all('p')
    for name1 in name:
        print(name1.get_text('title'))

    return name


def store_into_excel(filename, titleList, name, wb, ws1, n):
    r = 1
    if n is not 1:
        r = n + n - 1
    else:
        r = n
    # 设置表头
    # titleList = ['远光大数据平台']
    c = 1
    for row in range(len(titleList)):
        ws1.cell(row=r, column=c, value=titleList[row])
        c += 1
    c = 1
    # 填写表内容
    r += 1
    for name1 in name:
        ws1.cell(r, c, name1.encode('utf-8'))
        c += 1
    c = 1
    wb.save(filename=filename)

def main():
    # 将数据写入Excel
    wb = Workbook()
    # 设置Excel文件名
    filename = '大数据调研表2.xlsx'
    # 新建一个表
    ws1 = wb.active
    n = 1
    while True:
        print("\n\n\n**********进行第" + str(n) + "次爬取")
        print("网址：")
        url = Window(tk.Tk()).get_input()
        print(type(url))
        print("以获取到爬取连接： " + url)
        print("平台：")
        titleList = []
        titleList.append(Window(tk.Tk()).get_input())
        titleList.append(url)
        print("正在爬取数据")
        try:
            name = get_information(url)
            store_into_excel(filename, titleList, name, wb, ws1, n)
        except:
            print("输入有误。。。")
        print("输入任意值继续，输入q键退出。。")
        ans = Window(tk.Tk()).get_input()
        n += 1
        if ans is not 'q':
            continue
        else:
            break



if __name__ == "__main__":

    main()





