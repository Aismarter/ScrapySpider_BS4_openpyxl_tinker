import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook



url = "http://www.ygsoft.com/BigData/index.html?from=baidu"

page = requests.get(url)
print(page.status_code)
print(page.content)

soup = BeautifulSoup(page.content, 'html.parser')

print(soup.prettify())

name = soup.find_all('p')
for name1 in name:
    print(name1.get_text('title'))

# 将数据写入Excel
wb = Workbook()
# 设置Excel文件名
dest_filename = '工业大数据调研表.xlsx'
# 新建一个表
ws1 = wb.active

# 设置表头
titleList = ['远光大数据平台']
for row in range(len(titleList)):
    ws1.cell(row=1, column=1, value=titleList[row])

# 填写表内容
b =1
for name1 in name:
    ws1.cell(2, b, name1.encode('utf-8'))
    b += 1
wb.save(filename=dest_filename)






