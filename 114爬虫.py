import requests
import os
from tktest import *
from bs4 import BeautifulSoup
from openpyxl import Workbook
import openpyxl


def get_information(url):
    # url = "http://www.ygsoft.com/BigData/index.html?from=baidu"
    page = requests.get(url)
    # print(page.status_code)
    # print(page.content)
    soup = BeautifulSoup(page.content, 'html.parser')
    # print(soup.prettify())
    # details = soup.find_all('td', class_="text")
    # print(details.get_text())
    # return details
    name = soup.find_all('td', class_="lian-right")
    name_mess = []
    for name1 in name:
        mess = str(name1.get_text())
        print(mess)
        # print(name1.get_text())
        name_mess.append(mess)
    print(name_mess)
    return name_mess


def get_information_name(url):
    # url = "http://www.ygsoft.com/BigData/index.html?from=baidu"
    page = requests.get(url)
    # print(page.status_code)
    # print(page.content)
    soup = BeautifulSoup(page.content, 'html.parser')
    # print(soup.prettify())
    # details = soup.find_all('td', class_="text")
    # print(details.get_text())
    # return details
    name = soup.find_all('td', class_="top2")
    name_mess = []
    for name1 in name:
        mess = str(name1.get_text())
        print(mess)
        # print(name1.get_text())
        name_mess.append(mess)
    print(name_mess)
    return name_mess


def get_information_company(url):
    # url = "http://www.ygsoft.com/BigData/index.html?from=baidu"
    page = requests.get(url)
    # print(page.status_code)
    # print(page.content)
    soup = BeautifulSoup(page.content, 'html.parser')
    # print(soup.prettify())
    # details = soup.find_all('td', class_="text")
    # print(details.get_text())
    # return details
    name = soup.find_all('td', class_="text")
    name_mess = []
    for name1 in name:
        mess = str(name1.get_text())
        print(mess)
        # print(name1.get_text())
        name_mess.append(mess)
    print(name_mess)
    return name_mess


def store_into_excel(filename,titleList, company ,name, wb, ws1, n):
    r = 1
    if n is not 1:
        r = n + n -1
    else:
        r = n
    # 设置表头
    # titleList = ['远光大数据平台']
    c = 1
    for n1 in company:
        ws1.cell(r, c + 2, n1)
    for row in range(len(titleList)):
        ws1.cell(row=r, column=c, value=titleList[row])
        c += 1
    c = 1
    # 填写表内容
    r += 1
    for name1 in name:
        ws1.cell(r, c+2, name1)
        c += 1
    c = 1




    wb.save(filename=filename)

def main():
    # 将数据写入Excel
    wb = Workbook()
    # 设置Excel文件名
    filename = '黄页数据.xlsx'
    # 新建一个表
    ws1 = wb.active
    n = 1
    while True:
        print("\n\n\n**********进行第" + str(n) + "次爬取")
        print("网址：")
        url = Window(tk.Tk()).get_input()
        print(type(url))
        print("以获取到爬取连接： " + url)
        print("平台：")
        # titleList.append(Window(tk.Tk()).get_input())
        titleList = get_information_name(url)
        company = get_information_company(url)
        print("正在爬取数据")
        name = get_information(url)
        store_into_excel(filename, titleList, company, name, wb, ws1, n)
        # try:
        #     name = get_information(url)
        #     store_into_excel(filename,  name, wb, ws1, n)
        # except:
        #     print("输入有误。。。")
        print("输入任意值继续，输入q键退出。。")
        ans = Window(tk.Tk()).get_input()
        n += 1
        if ans is not 'q':
            continue
        else:
            break



if __name__ == "__main__":

    main()





